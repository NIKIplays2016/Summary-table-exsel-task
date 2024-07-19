import time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

from datetime import datetime


def save(pwd):
    wb.save(rf"{pwd}")


form22_dir = {
    "01" : "сельскохозяйственные организации, использующие предоставленные им земли для ведения сельского хозяйства, в том числе в исследовательских и учебных целях, а также для ведения подсобного хозяйства",
    "02" : "сельскохозяйственные организации Министерства сельского хозяйства и продовольствия Республики Беларусь",
    "03" : "крестьянские (фермерские) хозяйства",
    "05" : "граждане, использующие земельные участки для строительства и (или) обслуживания жилого дома",
    "06" : "граждане, использующие земельные участки для ведения личного подсобного хозяйства",
    "07" : "граждане, использующие земельные участки для садоводства и дачного строительства",
    "08" : "граждане, использующие земельные участки для огородничества",
    "09" : "граждане, использующие земельные участки для сенокошения и выпаса сельскохозяйственных животных",
    "10" : "граждане, использующие земельные участки для иных сельскохозяйственных целей",
    "11" : "граждане, использующие земельные участки для иных несельскохозяйственных целей",
    "12" : "промышленные организации",
    "13" : "организации железнодорожного транспорта",
    "14" : "организации автомобильного транспорта",
    "15" : "организации Вооруженных Сил Республики Беларусь, воинских частей, военных учебных заведений и других войск и воинских формирований Республики Беларусь",
    "16" : "организации воинских частей, военных учебных заведений и других войск и воинских формирований иностранных государств",
    "17" : "организации связи, энергетики, строительства, торговли, образования, здравоохранения и иные землепользователи",
    "18" : "организации природоохранного, оздоровительного, рекреационного и историко - культурного назначения",
    "19" : "заповедники, национальные парки и дендрологические парки",
    "20" : "организации, ведущие лесное хозяйство",
    "21" : "организации, эксплуатирующие и обслуживающие гидротехнические и иные водохозяйственные сооружения",
    "22" : "земли, земельные участки, не предоставленные землепользователям",
    "23" : "земли общего пользования в населенных пунктах, садоводческих товариществах и дачных кооперативах, а также земельные участки, используемые гражданами",
    "24" : "иные земли общего пользования за пределами границ населенных пунктов"

}


#функция форматирования таблицы
def format_size(x, width, y, height):
    """
    Функция расширяет ячейки и делает так, чтобы строки превышющие длину переносились
    Работает от A до ZY (x = от 0 до 575)
    """
    excel_column_letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
    first_letter = ""
    for a in range(x//24+1):
        for i in range(24):
            column = first_letter + excel_column_letters[i]
            ws.column_dimensions[column].width = width
        first_letter = excel_column_letters[a]

    for i in range(1, y+1):
        ws.row_dimensions[i].height = height


def create_excel():
    # Категории заголовков
    headers = [
        "Всего земель под древесно-кустарниковой растительностью по данным ЗИС",
        "подлежит включению в границы населенного пункта для его развития",
        "подлежит вовлечению в сельскохозяйственный оборот",
        "подлежит вовлечению в лесохозяйственный оборот",
        "подлежит вовлечению для использования в иных целях",
        "включено в границы населенного пункта для его развития",
        "вовлечено в сельскохозяйственный оборот",
        "вовлечено в лесохозяйственный оборот",
        "вовлечено для использования в иных целях",
        "не могут быть использованы в хозяйственной деятельности",
        "обследовано местным исполнительным комитетом",
        "не обследовано местным исполнительным комитетом"
    ]



    # Запись заголовков
    start_col = 5
    ws.row_dimensions[7].height = 50
    for i, header in enumerate(headers):
        ws.merge_cells(start_row=7, start_column=start_col, end_row=7, end_column=start_col + 1)
        ws.cell(row=7, column=start_col, value=header).alignment = Alignment(horizontal='center', vertical='center')
        start_col += 2


    sub_headers = ["количество контуров", "площадь, га"]
    # Запись подзаголовков
    start_col = 5
    for i in range(len(headers)*2):
        ws.cell(row=8, column=start_col, value=sub_headers[i % 2]).alignment = Alignment(horizontal='center', vertical='center')
        start_col += 1
    start_col = 1

    for i in range(2):
        ws.merge_cells(start_row=9, start_column=start_col, end_row=9, end_column=start_col + 1)
        ws.cell(row=9, column=start_col, value=i + 1).alignment = Alignment(horizontal='center', vertical='center')
        start_col += 2

    start_col=5
    for i in range(3,len(headers)*2-1):
        ws.cell(row=9, column=start_col, value=i).alignment = Alignment(horizontal='center', vertical='center')
        start_col += 1


def create_info_data(date1, date2, sity):
    ws['A3'] = 'Город: '
    ws['B3'] = f"{sity}"

    if date1 or date2:
        ws['A5'] = 'За период:'
        if date1:
            ws['B5'] = f"с {date1} "
        if date2:
            ws['C5'] = f"по {date2}"
    else:
        ws['A5'] = 'За весь период'

#('Витебская обл.', 'Докшицкий р-н', '02', 3, 1.4264)
def analyzing(rows:tuple) -> None:
    """ЖЕСТЬ"""
    pattern = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    xl_list = []

    sity_row = pattern.copy()
    sity_row[0] = rows[0][0]
    xl_list.append(sity_row)

    #new_row = pattern.copy()
    rayons = []
    for i in rows:
        rayons.append(i[1])
    rayons = sorted(set(rayons))

    copy_rows = list(rows)
    form_dir = {}       #ловарь хранящий формы для каждого района

    for rayon in rayons:
        delete_ind = []
        form_dir[rayon] = []

        for i, row in enumerate(copy_rows):
            if row[1] == rayon:
                form_dir[rayon].append(row[2])
                delete_ind.append(i)

        for i in reversed(delete_ind):    #Удаление какашек, чтоб в следующих итерациях не воняли
            del copy_rows[i]

        form_dir[rayon] = sorted(set(form_dir[rayon]))


    vovlevh_column = {
        1:6,
        2:8,
        3:10,
        4:12,
        5:22,
        6:14,
        7:16,
        8:18,
        9:20,
        0:26,
    }

    plus_collumn = 3

    copy_rows = list(rows)
    for rayon in rayons:
        rayon_block = [pattern.copy()]
        rayon_block[0][0] = rayon


        for form in form_dir[rayon]:
            form_row = pattern.copy()
            delete_ind = []
            form_row[2] = f"{form} - {form22_dir[form]}"

            for i, row in enumerate(copy_rows):
                form_column = vovlevh_column[row[3]]

                if row[2] == form and row[1] == rayon:
                    form_row[form_column + 1] += row[4]
                    form_row[form_column] += 1
                    delete_ind.append(i)


            sum_data = form_row[5:25]
            contur_sum = 0
            area_sum = 0

            for i, value in enumerate(sum_data):
                if i % 2 == 0:  # Четные индексы
                    area_sum += value
                else:  # Нечетные индексы
                    contur_sum += value

            form_row[24], form_row[25] = contur_sum, area_sum
            form_row[4], form_row[5] = contur_sum + form_row[24], area_sum + form_row[25]
            rayon_block.append(form_row)

            for i in reversed(delete_ind):
                del copy_rows[i]

        for row in rayon_block[1:]:
            for i in range(4, 28):
                rayon_block[0][i] += row[i]
                xl_list[0][i] += row[i]
        xl_list += rayon_block

    start_row = 10
    start_col = 1

    for row_idx, row_data in enumerate(xl_list, start=start_row):
        for col_idx, value in enumerate(row_data, start=start_col):
            if value == 0:
                ws.cell(row=row_idx, column=col_idx, value="")
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)



def xl_main(start_time: str, end_time: str, rows: tuple, pwd: str):
    start_time_test = time.time()
    global wb
    global ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводная таблица"

    try:
        sity = rows[0][0]
    except IndexError:
        raise TabError

    create_info_data(start_time, end_time, sity)
    format_size(50, 10, 9, 30)
    create_excel()
    analyzing(rows)

    border_thick = Side(border_style="medium", color="000000")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            cell.font = Font(size=9, bold=True)
            cell.border = Border(left=border_thick, right=border_thick, top=border_thick, bottom=border_thick, )

    thin_border = Side(border_style="thin", color="000000")  # Тонкая черная граница
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        # Применяем стиль границы к диапазону ячеек от A1 до AB6
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=28):  # А1 до AB6
        for cell in row:
            cell.border = border

    last_row = ws.max_row
    for row in range(10, last_row + 1):
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)  # C=3, D=4
        merged_cell = ws.cell(row=row, column=3)
        merged_cell.alignment = merged_cell.alignment.copy(wrap_text=False)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    ws.merge_cells('A7:B8')
    ws['A7'] = "Наименование административно-территориальной единицы (район, город областного подчинения)"
    ws.merge_cells('C7:D8')
    ws['C7'] = "Категория землепользователя по Форме 22"

    ws.merge_cells('A1:Z1')
    ws['A1'] = 'ИНФОРМАЦИЯ О ВОВЛЕЧЕНИИ ЗЕМЕЛЬ ПОД ДРЕВЕСНО-КУСТАРНИКОВОЙ РАСТИТЕЛЬНОСТЬЮ В ХОЗЯЙСТВЕННЫЙ ОБОРОТ'
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].font = Font(size=14)

    end_time_test = time.time()
    execution_time = end_time_test - start_time_test
    print(execution_time)

    save(pwd)





